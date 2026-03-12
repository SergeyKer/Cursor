import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_DIR = Path(__file__).parent
EXCEL_PATH = ROOT_DIR / "CS_процессы_и_скрипты.xlsx"
DATA_DIR = ROOT_DIR / "data"

MENU_GROUPS = {
    "incoming_operator": "1. Входящие звонки (ОПЕРАТОР):",
    "incoming_manager": "2. Входящие звонки (КЛИЕНТСКИЙ МЕНЕДЖЕР):",
    "outgoing_manager": "3. Исходящие звонки (КЛИЕНТСКИЙ МЕНЕДЖЕР):",
    "accounting": "4. Бухгалтерия:",
    "other": "Прочее",
}


def _normalize_section_title(s: str) -> str:
    """Нормализовать заголовок раздела из Excel для сопоставления."""
    low = (s or "").strip().lower()
    # убираем ведущие номера вида "1.", "2)" и т.п.
    while low and (low[0].isdigit() or low[0] in ". )\t"):
        low = low[1:]
    low = low.strip()
    return " ".join(low.split())


def _infer_menu_group_from_heading(text: str) -> Optional[str]:
    """
    Определить menu_group по строке-заголовку из Excel.
    Ожидаемые заголовки: входящие/исходящие звонки и бухгалтерия (часто с двоеточием).
    """
    low = _normalize_section_title(text)
    if not low:
        return None
    if "бухгалтер" in low:
        return "accounting"
    if "входящ" in low and "оператор" in low:
        return "incoming_operator"
    if "входящ" in low and "клиентск" in low and "менедж" in low:
        return "incoming_manager"
    if "исходящ" in low and "клиентск" in low and "менедж" in low:
        return "outgoing_manager"
    return None


def find_process_list_sheet(wb) -> Worksheet:
    """
    Try to find the main sheet that contains the list of processes.
    Heuristics: exact name 'Список процессов' or any sheet containing this phrase.
    """
    for name in wb.sheetnames:
        if name.strip().lower() == "список процессов":
            return wb[name]
    for name in wb.sheetnames:
        if "список" in name.lower() and "процесс" in name.lower():
            return wb[name]
    # Fallback: first sheet
    return wb[wb.sheetnames[0]]


def is_process_sheet_name(name: str) -> bool:
    """
    Process sheets start with 'Описание процесса' or are not the main list/tools.
    """
    low = name.strip().lower()
    if low.startswith("описание процесса"):
        return True
    if low.startswith("список процессов"):
        return False
    if "communication" in low and "tool" in low:
        return False
    # Любой другой лист считаем процессом (описание/карточка)
    return True


def parse_process_list(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Parse the main 'Список процессов' sheet into a list of processes.

    The exact column layout may differ from file to file, so we use a flexible
    approach: look at header row and map known column names if possible.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index = None
    for idx, row in enumerate(rows):
        joined = " ".join(str(c) for c in row if c is not None).lower()
        # Ищем строку с заголовками таблицы: "Название процесса" или "Номер" + название
        if "название процесса" in joined or ("номер" in joined and "название" in joined):
            header_row_index = idx
            break
    if header_row_index is None:
        for idx, row in enumerate(rows):
            joined = " ".join(str(c) for c in row if c is not None).lower()
            if "процесс" in joined and "этап" not in joined:
                header_row_index = idx
                break
    if header_row_index is None:
        header_row_index = 0

    headers = [str(c).strip() if c is not None else "" for c in rows[header_row_index]]

    def col_index(names: List[str]) -> Optional[int]:
        low_headers = [h.lower() for h in headers]
        for name in names:
            name_low = name.lower()
            if name_low in low_headers:
                return low_headers.index(name_low)
        return None

    idx_name = col_index(["Название процесса", "Процесс", "Название"])
    idx_desc = col_index(["Описание", "Краткое описание"])
    idx_sheet = col_index(["Лист", "Вкладка", "Sheet"])
    # Если колонки "Название процесса" нет, берём вторую колонку (часто "Название процесса" во 2-й)
    if idx_name is None and len(headers) >= 2:
        idx_name = 1

    processes: List[Dict[str, Any]] = []
    current_group: Optional[str] = None
    for row in rows[header_row_index + 1 :]:
        if all(cell is None for cell in row):
            continue

        name = row[idx_name] if idx_name is not None and idx_name < len(row) else None
        if not name:
            continue

        raw_name = str(name).strip()
        # Заголовки разделов в Excel: "1. Входящие звонки (ОПЕРАТОР):" и т.п.
        inferred = _infer_menu_group_from_heading(raw_name)
        if inferred:
            current_group = inferred
            continue

        description = (
            row[idx_desc] if idx_desc is not None and idx_desc < len(row) else None
        )
        sheet_name = (
            row[idx_sheet] if idx_sheet is not None and idx_sheet < len(row) else None
        )

        processes.append(
            {
                "code": raw_name,  # can be adjusted later
                "name": raw_name,
                "short_description": str(description).strip() if description else "",
                "sheet_name": str(sheet_name).strip() if sheet_name else "",
                "menu_group": current_group or "other",
            }
        )

    return processes


def find_header_row_for_stages(ws: Worksheet) -> Optional[int]:
    """
    Find the header row for the stages table by looking for known Russian headers.
    """
    target_headers = {
        "этап",
        "описание",
        "действия оператора",
        "срок (sla)",
        "цель",
        "рекомендация",
    }

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        values = [str(c).strip().lower() for c in row if c is not None]
        if not values:
            continue
        matched = sum(1 for v in values if v in target_headers)
        if matched >= 3:
            return idx
    return None


def parse_process_description(ws: Worksheet) -> Dict[str, str]:
    """
    Извлечь блок «Описание процесса» с листа: бизнес-цель и «Процесс позволяет».
    Ищем текст до таблицы этапов.
    """
    rows = list(ws.iter_rows(values_only=True))
    stages_row = find_header_row_for_stages(ws)
    if stages_row is None:
        stages_row = len(rows)
    lines: List[str] = []
    for row in rows[:stages_row]:
        parts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if parts:
            lines.append(" ".join(parts))
    full_text = "\n".join(lines)
    low = full_text.lower()
    idx_bg = low.find("бизнес-цель процесса")
    if idx_bg == -1:
        idx_bg = low.find("бизнес-цель")
    idx_pa = low.find("процесс позволяет")
    business_goal = ""
    process_allows = ""
    if idx_bg >= 0:
        colon = full_text.find(":", idx_bg)
        start = colon + 1 if colon >= 0 else idx_bg + 20
        end = idx_pa if idx_pa >= 0 else len(full_text)
        business_goal = full_text[start:end].strip()
    if idx_pa >= 0:
        colon = full_text.find(":", idx_pa)
        start = colon + 1 if colon >= 0 else idx_pa + 18
        process_allows = full_text[start:].strip()
    _strip_incoming_calls = lambda s: s.split("1. Входящие звонки (ОПЕРАТОР):")[0].strip() if s else s
    for stop in ("\nВВОДНЫЙ БЛОК", "\nОГЛАВЛЕНИЕ", "\n1. Входящие звонки"):
        pos = process_allows.find(stop)
        if pos >= 0:
            process_allows = process_allows[:pos].strip()
            break
    process_allows = _strip_incoming_calls(process_allows)
    for stop in ("\nВВОДНЫЙ БЛОК", "\nОГЛАВЛЕНИЕ", "\n1. Входящие звонки"):
        pos = business_goal.find(stop)
        if pos >= 0:
            business_goal = business_goal[:pos].strip()
            break
    business_goal = _strip_incoming_calls(business_goal)
    return {"business_goal": business_goal, "process_allows": process_allows}


def parse_stages(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Extract the stages table from a process sheet.
    """
    header_row_index = find_header_row_for_stages(ws)
    if header_row_index is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    headers_raw = rows[header_row_index]
    headers = [str(c).strip().lower() if c is not None else "" for c in headers_raw]

    def idx(name: str) -> Optional[int]:
        normalized = name.lower()
        for i, h in enumerate(headers):
            if h == normalized:
                return i
        return None

    idx_stage = idx("этап")
    idx_desc = idx("описание")
    idx_actions = idx("действия оператора")
    idx_sla = idx("срок (sla)")
    idx_goal = idx("цель")
    idx_recommendation = idx("рекомендация")

    stages: List[Dict[str, Any]] = []

    for row in rows[header_row_index + 1 :]:
        # stop when we hit a completely empty row
        if all(cell is None for cell in row):
            # allow small gaps; break at first long empty region
            if stages:
                break
            continue

        def get(col_idx: Optional[int]) -> str:
            if col_idx is None or col_idx >= len(row):
                return ""
            value = row[col_idx]
            return "" if value is None else str(value).strip()

        stage_name = get(idx_stage)
        description = get(idx_desc)
        actions = get(idx_actions)
        sla = get(idx_sla)
        goal = get(idx_goal)
        recommendation = get(idx_recommendation)

        # skip rows that have no meaningful data
        if not any([stage_name, description, actions, sla, goal, recommendation]):
            continue

        stages.append(
            {
                "stage": stage_name,
                "description": description,
                "operator_actions": actions,
                "sla": sla,
                "stage_goal": goal,
                "recommendations": recommendation,
            }
        )

    return stages


def extract_text_block(ws: Worksheet, start_row: int, end_row: int) -> str:
    """
    Concatenate cell values in the given row range into a plain text block.
    This is a generic helper which can be adjusted once exact layout is known.
    """
    lines: List[str] = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        # join non-empty cells with space
        parts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if parts:
            lines.append(" ".join(parts))
    return "\n".join(lines).strip()


def parse_scripts(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Parse script tables on the sheet.

    We support two main layouts:
      - Этап разговора | Формулировка (пример) | Цель
      - Жалоба клиента | Ответ оператора | Цель
    """
    rows = list(ws.iter_rows(values_only=True))
    scripts: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows):
        headers = [str(c).strip().lower() if c is not None else "" for c in row]
        if not any(headers):
            continue

        # Разговорный скрипт
        if "этап разговора" in headers and "формулировка" in headers:
            idx_step = headers.index("этап разговора")
            idx_phrase = headers.index("формулировка")
            idx_goal = headers.index("цель") if "цель" in headers else None

            for data_row in rows[idx + 1 :]:
                if all(c is None for c in data_row):
                    break
                step = (
                    str(data_row[idx_step]).strip()
                    if idx_step < len(data_row) and data_row[idx_step] is not None
                    else ""
                )
                phrase = (
                    str(data_row[idx_phrase]).strip()
                    if idx_phrase < len(data_row) and data_row[idx_phrase] is not None
                    else ""
                )
                goal = (
                    str(data_row[idx_goal]).strip()
                    if idx_goal is not None
                    and idx_goal < len(data_row)
                    and data_row[idx_goal] is not None
                    else ""
                )
                if not any([step, phrase, goal]):
                    continue
                scripts.append(
                    {
                        "type": "conversation",
                        "step_name": step,
                        "phrase": phrase,
                        "goal": goal,
                    }
                )

        # Жалобы клиента / ответы оператора (гибкий поиск столбцов)
        def find_col(pattern: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if pattern in h:
                    return i
            return None

        idx_complaint = find_col("жалоба клиента") or find_col("жалоба")
        idx_reply = find_col("ответ оператора") or find_col("ответ")
        idx_goal = find_col("цель")
        if idx_complaint is not None and idx_reply is not None:
            for data_row in rows[idx + 1 :]:
                if all(c is None for c in data_row):
                    break
                complaint = (
                    str(data_row[idx_complaint]).strip()
                    if idx_complaint < len(data_row)
                    and data_row[idx_complaint] is not None
                    else ""
                )
                reply = (
                    str(data_row[idx_reply]).strip()
                    if idx_reply < len(data_row) and data_row[idx_reply] is not None
                    else ""
                )
                goal = (
                    str(data_row[idx_goal]).strip()
                    if idx_goal is not None
                    and idx_goal < len(data_row)
                    and data_row[idx_goal] is not None
                    else ""
                )
                if not any([complaint, reply, goal]):
                    continue
                scripts.append(
                    {
                        "type": "complaint",
                        "complaint": complaint,
                        "reply": reply,
                        "goal": goal,
                    }
                )

    return scripts


def parse_difficult_phrases(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Parse table 'ФРАЗЫ ПРИ СЛОЖНЫХ СИТУАЦИЯХ':
    Тип клиента и сложной ситуации | Цель оператора | Готовые формулировки
    """
    rows = list(ws.iter_rows(values_only=True))
    result: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows):
        headers = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(headers)
        if "готовые формулировки" not in joined or "тип клиента" not in joined:
            continue

        idx_type = None
        idx_goal = None
        idx_phrases = None
        for i, h in enumerate(headers):
            if "тип клиента" in h and "сложн" in h:
                idx_type = i
            if "цель оператора" in h:
                idx_goal = i
            if "готовые формулировки" in h:
                idx_phrases = i
        if idx_phrases is None:
            continue
        if idx_type is None:
            idx_type = 0
        if idx_goal is None:
            idx_goal = 1 if idx_type == 0 else 0

        for data_row in rows[idx + 1 :]:
            if all(c is None for c in data_row):
                break
            raw_type = (
                str(data_row[idx_type]).strip()
                if idx_type < len(data_row) and data_row[idx_type] is not None
                else ""
            )
            raw_goal = (
                str(data_row[idx_goal]).strip()
                if idx_goal < len(data_row) and data_row[idx_goal] is not None
                else ""
            )
            raw_phrases = (
                str(data_row[idx_phrases]).strip()
                if idx_phrases < len(data_row) and data_row[idx_phrases] is not None
                else ""
            )
            if not raw_type and not raw_goal and not raw_phrases:
                continue
            phrases_list = [
                p.strip()
                for p in raw_phrases.replace("«", "").replace("»", "").split("\n")
                if p.strip()
            ]
            if not phrases_list and raw_phrases:
                phrases_list = [raw_phrases]
            result.append(
                {
                    "situation_type": raw_type,
                    "operator_goal": raw_goal,
                    "phrases": phrases_list if phrases_list else [raw_phrases] if raw_phrases else [],
                }
            )
        break

    return result


def _collect_text_after_title(rows: List[tuple], start_idx: int, stop_markers: List[str]) -> str:
    """Собрать текст из строк после start_idx до пустой строки или маркера."""
    lines: List[str] = []
    for row in rows[start_idx + 1 :]:
        if all(c is None for c in row):
            break
        row_str = " ".join(str(c).strip() for c in row if c is not None and str(c).strip())
        row_low = row_str.lower()
        if any(m in row_low for m in stop_markers):
            break
        if row_str:
            lines.append(row_str)
    return "\n".join(lines).strip()


def parse_email_and_auto_templates(ws: Worksheet) -> tuple:
    """
    Find blocks 'ШАБЛОН СООБЩЕНИЯ С ЭЛ. АДРЕСОМ' and 'ШАБЛОН АВТООТВЕТА ИЗ CRM'.
    Return (email_template_dict or None, auto_reply_template_dict or None).
    """
    rows = list(ws.iter_rows(values_only=True))
    email_tpl: Optional[Dict[str, Any]] = None
    auto_tpl: Optional[Dict[str, Any]] = None
    stop_markers = [
        "фразы при сложных",
        "шаблон сообщения",
        "шаблон автоответа",
        "шпаргалка",
        "оглавление",
    ]

    for idx, row in enumerate(rows):
        row_joined = " ".join(str(c) for c in row if c is not None).lower()
        if "шаблон сообщения" in row_joined and "эл" in row_joined:
            body = _collect_text_after_title(rows, idx, stop_markers)
            if body:
                email_tpl = {"subject": "", "body": body}
        if "шаблон автоответа" in row_joined and "crm" in row_joined:
            body = _collect_text_after_title(rows, idx, stop_markers)
            if body:
                auto_tpl = {"subject": "", "body": body}
                if "тема:" in body.lower():
                    parts = body.split("\n", 1)
                    if len(parts) >= 2 and "тема:" in parts[0].lower():
                        auto_tpl["subject"] = parts[0].replace("Тема:", "").replace("тема:", "").strip()
                        auto_tpl["body"] = parts[1].strip()

    return (email_tpl, auto_tpl)


def parse_cheatsheet(ws: Worksheet) -> List[str]:
    """Найти блок ШПАРГАЛКА и собрать список фраз."""
    rows = list(ws.iter_rows(values_only=True))
    phrases: List[str] = []

    for idx, row in enumerate(rows):
        row_joined = " ".join(str(c) for c in row if c is not None).lower()
        if "шпаргалка" not in row_joined:
            continue
        for data_row in rows[idx + 1 :]:
            if all(c is None for c in data_row):
                break
            for cell in data_row:
                if cell is None:
                    continue
                s = str(cell).strip()
                if not s or len(s) < 10:
                    continue
                if s.lower().startswith(("шаблон", "описание", "оглавление", "скрипт")):
                    break
                if s.startswith("«") or "спасибо" in s.lower() or "понят" in s.lower() or "вышлю" in s.lower():
                    phrases.append(s)
        break

    return phrases


def parse_process_sheet(name: str, ws: Worksheet) -> Dict[str, Any]:
    """
    Parse a single 'Описание процесса ...' sheet into a structured dict.

    Because we don't know exact cell coordinates for description, scripts and phrases,
    this function currently focuses on:
      - basic process metadata (from sheet title),
      - stages table (reliably detected by headers).

    Other blocks are left as empty strings/lists so they can be filled later or
    доработаны после уточнения формата файла.
    """
    process_description = parse_process_description(ws)
    stages = parse_stages(ws)
    scripts = parse_scripts(ws)
    difficult_phrases = parse_difficult_phrases(ws)
    email_tpl, auto_tpl = parse_email_and_auto_templates(ws)
    cheatsheet = parse_cheatsheet(ws)

    process_data: Dict[str, Any] = {
        "sheet_name": name,
        "name": name,
        "goal": process_description.get("business_goal", ""),
        "description": process_description.get("process_allows", ""),
        "process_description": process_description,
        "stages": stages,
        "script_steps": scripts,
        "main_script": "",
        "difficult_phrases": difficult_phrases,
        "email_template": email_tpl,
        "auto_reply_template": auto_tpl,
        "cheatsheet": cheatsheet,
    }

    return process_data


def parse_communication_tools(wb) -> List[Dict[str, Any]]:
    """
    Parse a sheet with general communication tools, if it exists.
    Ищем лист по названию: "Инструменты общения", "Инструменты коммуникации"
    или "communication tools".
    """
    sheet_name = None
    for name in wb.sheetnames:
        low = name.strip().lower()
        if "communication" in low and "tool" in low:
            sheet_name = name
            break
        if "инструмент" in low and "общени" in low:
            sheet_name = name
            break
        if "инструмент" in low and "коммуникац" in low:
            sheet_name = name
            break
    if not sheet_name:
        return []

    ws: Worksheet = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    blocks: List[Dict[str, Any]] = []
    current_title = ""
    current_body: List[str] = []

    for row in rows:
        if not row or all(cell is None for cell in row):
            if current_title or current_body:
                body = "\n".join(current_body).strip()
                if current_title or body:
                    blocks.append({"title": current_title or "Блок", "body": body})
            current_title = ""
            current_body = []
            continue
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        first = cells[0]
        if len(first) > 2 and (first[0].isupper() or first.startswith("•") or first.isdigit()):
            if current_title or current_body:
                body = "\n".join(current_body).strip()
                if current_title or body:
                    blocks.append({"title": current_title or "Блок", "body": body})
            current_title = first.lstrip("0123456789. )•-")
            current_body = []
            rest = cells[1:] if len(cells) > 1 else []
            if rest:
                current_body.append("\n".join(rest))
        else:
            current_body.append(" ".join(cells))

    if current_title or current_body:
        body = "\n".join(current_body).strip()
        if current_title or body:
            blocks.append({"title": current_title or "Блок", "body": body})

    return blocks


def migrate() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found at {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    DATA_DIR.mkdir(exist_ok=True)

    process_list_ws = find_process_list_sheet(wb)
    processes_meta = parse_process_list(process_list_ws)

    # Если из таблицы ничего не вытащили — собираем список из названий листов
    if not processes_meta:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower().startswith("список процессов"):
                continue
            low = sheet_name.strip().lower()
            if "communication" in low and "tool" in low:
                continue
            processes_meta.append({
                "code": sheet_name[:80],
                "name": sheet_name,
                "short_description": "",
                "sheet_name": sheet_name,
            })

    processes_data: Dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        if not is_process_sheet_name(sheet_name):
            continue
        ws = wb[sheet_name]
        process_struct = parse_process_sheet(sheet_name, ws)
        processes_data[sheet_name] = process_struct

    # Добавляем в меню процессы из листов, которых нет в «Список процессов»
    meta_names = {str(m.get("sheet_name") or m.get("name") or m.get("code") or "").strip() for m in processes_meta}
    for sheet_name in processes_data:
        if sheet_name.strip() and sheet_name not in meta_names:
            processes_meta.append({
                "code": sheet_name[:80],
                "name": sheet_name,
                "short_description": "",
                "sheet_name": sheet_name,
            })
            meta_names.add(sheet_name)

    communication_tools = parse_communication_tools(wb)

    def find_process_for_meta(meta: Dict[str, Any]) -> tuple:
        """Найти (ключ_листа, данные) для meta: точное совпадение или по вхождению имени."""
        sheet_name = (meta.get("sheet_name") or meta.get("name") or meta.get("code") or "").strip()
        if not sheet_name:
            return None, None
        if sheet_name in processes_data:
            return sheet_name, processes_data[sheet_name]
        name_lower = (meta.get("name") or sheet_name).lower()
        for key in processes_data:
            if name_lower in key.lower() or key.lower() in name_lower:
                return key, processes_data[key]
        return None, None

    # Добавляем searchable_text в meta для поиска по содержимому
    for meta in processes_meta:
        _, proc = find_process_for_meta(meta)
        parts = [meta.get("name", ""), meta.get("short_description", "")]
        if proc:
            pd = proc.get("process_description") or {}
            parts.append(pd.get("business_goal", "") or proc.get("goal", ""))
            parts.append(pd.get("process_allows", "") or proc.get("description", ""))
            for s in proc.get("stages") or []:
                parts.append(s.get("stage", ""))
                parts.append(s.get("description", ""))
            for step in proc.get("script_steps") or []:
                parts.append(step.get("step_name", "") or step.get("phrase", "") or step.get("reply", ""))
                parts.append(step.get("complaint", ""))
            for df in proc.get("difficult_phrases") or []:
                parts.append(df.get("situation_type", ""))
                for p in df.get("phrases") or []:
                    parts.append(p)
            for phrase in proc.get("cheatsheet") or []:
                parts.append(phrase)
        meta["searchable_text"] = " ".join(str(p) for p in parts if p).strip()

    # База знаний для ИИ: плоский список процессов с ключевыми текстами
    knowledge: List[Dict[str, Any]] = []
    for meta in processes_meta:
        sheet_name = meta.get("sheet_name") or meta.get("name")
        _, proc = find_process_for_meta(meta)
        if not proc:
            knowledge.append({"name": meta.get("name"), "sheet_name": sheet_name, "goal": "", "stages": [], "script_phrases": [], "difficult_phrases": [], "cheatsheet": []})
            continue
        stages_brief = [{"stage": s.get("stage"), "description": s.get("description")} for s in (proc.get("stages") or [])]
        script_phrases = []
        for step in proc.get("script_steps") or []:
            script_phrases.append(step.get("phrase") or step.get("reply") or step.get("step_name", ""))
        difficult = [{"situation": d.get("situation_type"), "phrases": d.get("phrases", [])} for d in (proc.get("difficult_phrases") or [])]
        knowledge.append({
            "name": proc.get("name") or meta.get("name"),
            "sheet_name": sheet_name,
            "goal": proc.get("goal", ""),
            "description": proc.get("description", ""),
            "stages": stages_brief,
            "script_phrases": script_phrases,
            "difficult_phrases": difficult,
            "cheatsheet": proc.get("cheatsheet", []),
            "email_template": proc.get("email_template"),
            "auto_reply_template": proc.get("auto_reply_template"),
        })

    # processes.json с ключами по имени из списка процессов (чтобы «Ответ оператора» и др. всегда находились)
    output_processes: Dict[str, Any] = {}
    matched_data_keys: set = set()
    for meta in processes_meta:
        canonical_name = (meta.get("name") or meta.get("sheet_name") or meta.get("code") or "").strip()
        data_key, proc = find_process_for_meta(meta)
        if proc is not None and canonical_name:
            output_processes[canonical_name] = proc
            if data_key:
                matched_data_keys.add(data_key)
    for sheet_name, proc in processes_data.items():
        if sheet_name not in matched_data_keys:
            output_processes[sheet_name] = proc

    # Write JSON files
    (DATA_DIR / "processes_meta.json").write_text(
        json.dumps(processes_meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "processes.json").write_text(
        json.dumps(output_processes, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "communication_tools.json").write_text(
        json.dumps(communication_tools, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "knowledge.json").write_text(
        json.dumps(knowledge, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Migration finished.")
    print(f"- processes_meta.json: {len(processes_meta)} процессов")
    print(f"- processes.json: {len(processes_data)} листов 'Описание процесса'")
    print(f"- communication_tools.json: {len(communication_tools)} блоков")
    print(f"- knowledge.json: {len(knowledge)} записей для ИИ")


if __name__ == "__main__":
    migrate()

